# -*- coding: utf-8 -*- 

"""
Author: Anger36
Date: 19-5-14
Describe: openpyxl功能扩展, 版本基于2.4.8
"""

from openpyxl import Workbook
from openpyxl.utils.exceptions import ReadOnlyWorkbookException
from openpyxl.writer.write_only import WriteOnlyWorksheet
from openpyxl.worksheet import Worksheet
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

__all__ = ['ReWorkbook', 'HEAD_STYLE', 'YELLOW_TITLE_STYLE', 'GREEN_TITLE_STYLE', 'ORANGE_TITLE_STYLE',
           'GRAY_TITLE_STYLE', 'ORDER_CENTENT_STYLE', 'OA_CONTENT_STYLE', 'OA_TITLE_STYLE', 'ABNORMAL_STYLE']


class ReWorksheet(Worksheet):
    DEFUALT_STYLE = ['border', 'font', 'fill', 'alignment']

    def __init__(self, parent, title=None):
        super(ReWorksheet, self).__init__(parent, title)
        self.expected_type = [str, int]

    def write_cell(self, row, column, value=None, style=None):
        self.cell(row=row, column=column, value=value)
        if style is not None:
            for st in self.DEFUALT_STYLE:
                if getattr(style, st) is not None:
                    setattr(self.cell(row=row, column=column), st, getattr(style, st))

    def write_merge(self, start_row, end_row, start_column, end_column, content, style=None):
        self.merge_cells(start_row=start_row, start_column=start_column, end_row=end_row, end_column=end_column)
        self.write_cell(start_row, start_column, content, style=style)

    def write_row(self, start_row, end_row, start_col, contents, style=None):
        if start_row == end_row:
            for content in contents:
                self.write_cell(start_row, start_col, content, style)
                start_col += 1
        else:
            for content in contents:
                self.write_merge(start_row, end_row, start_col, start_col, content, style)
                start_col += 1

    def set_row(self, row, height):
        self.row_dimensions[row].height = height

    def set_col(self, col, width):
        if type(col) == str:
            self.column_dimensions[col].width = width
        elif type(col) == int:
            col = get_column_letter(col)
            self.column_dimensions[col].width = width
        else:
            raise TypeError('expected column type is ' + str(self.expected_type))


class ReWorkbook(Workbook):
    def __init__(self):
        super(ReWorkbook, self).__init__()
        if not self.write_only:
            self._origin_sheet = self._sheets.pop(0)
            self._sheets.append(ReWorksheet(self))

    def create_sheet(self, title=None, index=None):
        """Create a worksheet (at an optional index).

        :param title: optional title of the sheet
        :type title: unicode
        :param index: optional position at which the sheet will be inserted
        :type index: int

        """
        if self.read_only:
            raise ReadOnlyWorkbookException('Cannot create new sheet in a read-only workbook')

        if self.write_only:
            new_ws = WriteOnlyWorksheet(parent=self, title=title)
        else:
            new_ws = ReWorksheet(parent=self, title=title)

        self._add_sheet(sheet=new_ws, index=index)
        return new_ws

    @property
    def active(self):
        """Get the currently active sheet or None

        :type: :class:`ReWorksheet`
        """
        try:
            return self._sheets[self._active_sheet_index]
        except IndexError:
            pass


HEAD_FONT = Font(name=u'楷体', sz=28, b=True)
YELLOW_BACK_COLOR = PatternFill(fill_type='solid', fgColor="00FFFF00")
GREEN_BACK_COLOR = PatternFill(fill_type='solid', fgColor="0092D050")
GRAY_BACK_COLOR = PatternFill(fill_type='solid', fgColor="00DEDEDE")
ORANGE_BACK_COLOR = PatternFill(fill_type='solid', fgColor="00F89450")

TITLE_FONT = Font(sz=14)
RED_FONT = Font(color="00FF0000")

BASIC_BORDER = Border(left=Side(border_style='thin', color='000000'),
                      right=Side(border_style='thin', color='000000'),
                      top=Side(border_style='thin', color='000000'),
                      bottom=Side(border_style='thin', color='000000'))


class OpenpyxlStyle(object):
    def __init__(self, font=None, patternfill=None, border=None, alignment=None):
        # self.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.alignment = alignment
        self.border = border
        self.font = font
        self.fill = patternfill


ALIGNMENT1 = Alignment(vertical='center')
ALIGNMENT2 = Alignment(horizontal='center', vertical='center')

# order report style
ORDER_CENTENT_STYLE = OpenpyxlStyle(alignment=ALIGNMENT1)
HEAD_STYLE = OpenpyxlStyle(font=HEAD_FONT, alignment=ALIGNMENT2)
YELLOW_TITLE_STYLE = OpenpyxlStyle(font=TITLE_FONT, patternfill=YELLOW_BACK_COLOR, alignment=ALIGNMENT2)
GREEN_TITLE_STYLE = OpenpyxlStyle(font=TITLE_FONT, patternfill=GREEN_BACK_COLOR, alignment=ALIGNMENT2)
ORANGE_TITLE_STYLE = OpenpyxlStyle(font=TITLE_FONT, patternfill=ORANGE_BACK_COLOR, alignment=ALIGNMENT2)
GRAY_TITLE_STYLE = OpenpyxlStyle(font=TITLE_FONT, patternfill=GRAY_BACK_COLOR, border=BASIC_BORDER, alignment=ALIGNMENT1)
ABNORMAL_STYLE = OpenpyxlStyle(font=RED_FONT)

# OA report style
OA_CONTENT_STYLE = OpenpyxlStyle(alignment=ALIGNMENT2, border=BASIC_BORDER)
OA_TITLE_STYLE = OpenpyxlStyle(font=Font(sz=18, name=u'楷体', b=True), patternfill=GREEN_BACK_COLOR,
                               border=BASIC_BORDER, alignment=ALIGNMENT2)
